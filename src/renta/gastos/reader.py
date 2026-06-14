"""Read gastos deducibles from the Excel template."""

from __future__ import annotations

import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

from renta.gastos.categories import Beneficiary, DeclarationType, DeductionCategory, PaymentMethod
from renta.gastos.models import ContributorConfig, ExpenseEntry
from renta.utils.exceptions import ParseError

_EXPENSE_COLUMNS = (
    "fecha",
    "categoria",
    "descripcion",
    "proveedor",
    "nif_proveedor",
    "importe",
    "metodo_pago",
    "beneficiario",
    "tiene_factura",
    "tiene_justificante_pago",
    "notas",
)

_CONFIG_KEYS = (
    "base_liquidable_general",
    "base_liquidable_ahorro",
    "tipo_declaracion",
    "edad_contribuyente",
    "grado_discapacidad",
    "tiene_familia_numerosa",
    "contribuyente_en_paro",
)


def read_gastos_excel(path: Path) -> tuple[ContributorConfig, list[ExpenseEntry]]:
    """Read contributor config and expense entries from the Excel template.

    Raises ParseError on missing sheets, missing columns, or invalid values.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    config = _read_config(wb, path)
    entries = _read_expenses(wb, path)
    return config, entries


def _read_config(wb: openpyxl.Workbook, path: Path) -> ContributorConfig:
    if "config" not in wb.sheetnames:
        raise ParseError(f"{path.name}: missing sheet 'config'")
    ws = wb["config"]
    data: dict[str, object] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1] is not None:
            data[str(row[0]).strip()] = row[1]

    missing = [k for k in _CONFIG_KEYS if k not in data]
    if missing:
        raise ParseError(f"{path.name} config sheet: missing keys: {missing}")

    try:
        return ContributorConfig(
            base_liquidable_general=_parse_decimal(
                data["base_liquidable_general"], "base_liquidable_general"
            ),
            base_liquidable_ahorro=_parse_decimal(
                data["base_liquidable_ahorro"], "base_liquidable_ahorro"
            ),
            declaration_type=DeclarationType(str(data["tipo_declaracion"]).strip().lower()),
            age=int(str(data["edad_contribuyente"])),
            disability_pct=int(str(data.get("grado_discapacidad") or 0)),
            has_familia_numerosa=_parse_bool(
                data.get("tiene_familia_numerosa"), "tiene_familia_numerosa"
            ),
            contribuyente_en_paro=_parse_bool(
                data.get("contribuyente_en_paro"), "contribuyente_en_paro"
            ),
        )
    except ValueError as exc:
        raise ParseError(f"{path.name} config sheet: {exc}") from exc


def _read_expenses(wb: openpyxl.Workbook, path: Path) -> list[ExpenseEntry]:
    if "gastos" not in wb.sheetnames:
        raise ParseError(f"{path.name}: missing sheet 'gastos'")
    ws = wb["gastos"]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise ParseError(f"{path.name} gastos sheet: empty header row")

    columns = {str(cell).strip().lower(): idx for idx, cell in enumerate(header_row) if cell}
    missing = [c for c in _EXPENSE_COLUMNS if c not in columns]
    if missing:
        raise ParseError(f"{path.name} gastos sheet: missing columns: {missing}")

    entries: list[ExpenseEntry] = []
    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        amount_raw = row[columns["importe"]]
        if amount_raw is None or str(amount_raw).strip() == "":
            continue
        try:
            amount = _parse_decimal(amount_raw, f"row {row_num} importe")
        except ParseError:
            continue
        if amount <= Decimal("0"):
            continue

        try:
            entry = _parse_row(row, columns, row_num)
        except ParseError as exc:
            raise ParseError(f"{path.name} gastos row {row_num}: {exc}") from exc
        entries.append(entry)

    return entries


def _parse_row(row: tuple[object, ...], columns: dict[str, int], row_num: int) -> ExpenseEntry:
    def get(col: str) -> object:
        return row[columns[col]]

    fecha_raw = get("fecha")
    if isinstance(fecha_raw, datetime.datetime):
        date = fecha_raw.date()
    elif isinstance(fecha_raw, datetime.date):
        date = fecha_raw
    elif isinstance(fecha_raw, str):
        try:
            date = datetime.date.fromisoformat(fecha_raw.strip())
        except ValueError as exc:
            raise ParseError(f"invalid date: {fecha_raw!r}") from exc
    else:
        raise ParseError(f"invalid date type: {type(fecha_raw)}")

    try:
        category = DeductionCategory(str(get("categoria")).strip().lower())
    except ValueError as exc:
        raise ParseError(f"unknown category: {get('categoria')!r}") from exc

    try:
        payment_method = PaymentMethod(str(get("metodo_pago")).strip().lower())
    except ValueError as exc:
        raise ParseError(f"unknown payment_method: {get('metodo_pago')!r}") from exc

    try:
        beneficiary = Beneficiary(str(get("beneficiario")).strip().lower())
    except ValueError as exc:
        raise ParseError(f"unknown beneficiary: {get('beneficiario')!r}") from exc

    nif = str(get("nif_proveedor") or "").strip()
    if not nif:
        raise ParseError("nif_proveedor is required")

    return ExpenseEntry(
        date=date,
        category=category,
        description=str(get("descripcion") or "").strip(),
        provider=str(get("proveedor") or "").strip(),
        provider_nif=nif,
        amount=_parse_decimal(get("importe"), f"importe row {row_num}"),
        payment_method=payment_method,
        beneficiary=beneficiary,
        has_invoice=_parse_bool(get("tiene_factura"), "tiene_factura"),
        has_payment_proof=_parse_bool(get("tiene_justificante_pago"), "tiene_justificante_pago"),
        notes=str(get("notas") or "").strip(),
    )


def _parse_decimal(value: object, field: str) -> Decimal:
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, Decimal):
        return value
    s = str(value).strip().replace(",", ".")
    try:
        return Decimal(s)
    except InvalidOperation as exc:
        raise ParseError(f"cannot parse decimal for {field}: {value!r}") from exc


def _parse_bool(value: object, field: str) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        return value.strip().upper() in ("TRUE", "SI", "SÍ", "1", "YES")
    raise ParseError(f"cannot parse boolean for {field}: {value!r}")
