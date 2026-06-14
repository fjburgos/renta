"""Generate the gastos_deducibles Excel template."""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from renta.gastos.categories import (
    Beneficiary,
    DeclarationType,
    DeductionCategory,
    PaymentMethod,
)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
_SECTION_FONT = Font(bold=True)
_NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")

_EXPENSE_HEADERS = [
    ("fecha", 14),
    ("categoria", 22),
    ("descripcion", 35),
    ("proveedor", 25),
    ("nif_proveedor", 14),
    ("importe", 12),
    ("metodo_pago", 16),
    ("beneficiario", 16),
    ("tiene_factura", 14),
    ("tiene_justificante_pago", 22),
    ("notas", 40),
]


def create_template(output_path: Path, year: int = 2025) -> None:
    """Write an empty gastos_deducibles template Excel file."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # type: ignore[arg-type]

    _build_config_sheet(wb)
    _build_gastos_sheet(wb)
    _build_instructions_sheet(wb, year)

    wb.save(output_path)


def _build_config_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("config")
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    _write_header(ws, "A1", "Parámetro")
    _write_header(ws, "B1", "Valor")

    rows = [
        ("base_liquidable_general", 0),
        ("base_liquidable_ahorro", 0),
        ("tipo_declaracion", "individual"),
        ("edad_contribuyente", 0),
        ("grado_discapacidad", 0),
        ("tiene_familia_numerosa", False),
        ("contribuyente_en_paro", False),
    ]
    for i, (key, default) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=key)
        ws.cell(row=i, column=2, value=default)

    # Dropdown for tipo_declaracion
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(d.value for d in DeclarationType)}"',
        allow_blank=False,
    )
    ws.add_data_validation(dv)
    dv.add("B4")


def _build_gastos_sheet(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("gastos")

    # Headers
    for col, (name, width) in enumerate(_EXPENSE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"

    # Data validations (applied to rows 2..500)
    _add_dropdown(ws, col=2, values=[c.value for c in DeductionCategory], rows=500)
    _add_dropdown(ws, col=7, values=[p.value for p in PaymentMethod], rows=500)
    _add_dropdown(ws, col=8, values=[b.value for b in Beneficiary], rows=500)
    _add_dropdown(ws, col=9, values=["TRUE", "FALSE"], rows=500)
    _add_dropdown(ws, col=10, values=["TRUE", "FALSE"], rows=500)

    # Date format for column 1
    for row in range(2, 502):
        ws.cell(row=row, column=1).number_format = "YYYY-MM-DD"
        ws.cell(row=row, column=6).number_format = "#,##0.00"

    # Example row (greyed out as a hint)
    example = [
        "2025-01-15",
        "deporte_saludable",
        "Cuota mensual gimnasio enero",
        "Gimnasio Ejemplo S.L.",
        "B12345678",
        "40.00",
        "tarjeta",
        "contribuyente",
        "TRUE",
        "TRUE",
        "",
    ]
    for col, val in enumerate(example, start=1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.fill = _NOTE_FILL


def _build_instructions_sheet(wb: openpyxl.Workbook, year: int) -> None:
    ws = wb.create_sheet("instrucciones")
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 22

    title = ws.cell(row=1, column=1, value=f"Gastos Deducibles CV — IRPF {year}")
    title.font = Font(bold=True, size=13)
    ws.merge_cells("A1:D1")

    _write_header(ws, "A3", "Categoría")
    _write_header(ws, "B3", "Descripción")
    _write_header(ws, "C3", "% / regla")
    _write_header(ws, "D3", "Límite anual")

    categories = [
        ("deporte_saludable", "Gimnasio, club deportivo, entrenador, fisio, dietista",
         "30% (50%/100% edad/discap.)", "150 €"),
        ("salud_bucodental", "Dentista (no estético)", "30%", "150 €"),
        ("salud_mental", "Psicólogo clínico, psiquiatra", "30%", "150 €"),
        ("optica", "Gafas graduadas, lentes de contacto", "30%", "100 €"),
        ("enfermedad_cronica", "Enf. crónicas complejas, raras, Alzheimer",
         "hasta 100 € fijo", "100 €"),
        ("formacion_musical", "Conservatorio, escuela música inscrita CV", "100%", "150 €"),
        ("abonos_culturales", "Abonos 'Abono Cultural Valenciano' (Culturarts)",
         "21%", "base 165 €"),
        ("guarderia", "Guardería / 1er ciclo infantil (< 3 años)", "15%", "297 €/menor"),
        ("material_escolar", "Material escolar (Primaria/ESO, solo desempleados)",
         "fijo 110 €/hijo", "110 €/hijo"),
    ]
    for i, (cat, desc, rule, limit) in enumerate(categories, start=4):
        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=2, value=desc)
        ws.cell(row=i, column=3, value=rule)
        ws.cell(row=i, column=4, value=limit)

    ws.cell(row=14, column=1, value="Notas importantes").font = _SECTION_FONT
    notes = [
        "El pago en efectivo NO es válido para ninguna deducción.",
        "Conserva: factura (o recibo con NIF) + justificante de pago (tarjeta/transferencia).",
        "Los 4 conceptos de salud (bucodental, mental, óptica, enf. crónicas) son acumulables.",
        "Límite de renta individual: 60.000 € (reducción proporcional 54.000–60.000 €).",
        "Guardería y material escolar: límite de renta más bajo (30.000 € individual).",
    ]
    for i, note in enumerate(notes, start=15):
        ws.cell(row=i, column=1, value=f"• {note}").fill = _NOTE_FILL
        ws.merge_cells(f"A{i}:D{i}")


def _write_header(ws: Worksheet, cell_ref: str, value: str) -> None:
    cell = ws[cell_ref]
    cell.value = value
    cell.fill = _SECTION_FILL
    cell.font = _SECTION_FONT


def _add_dropdown(
    ws: Worksheet,
    col: int,
    values: list[str],
    rows: int,
) -> None:
    formula = '"' + ",".join(values) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    col_letter = get_column_letter(col)
    dv.add(f"{col_letter}2:{col_letter}{rows + 1}")


if __name__ == "__main__":
    import sys

    year = int(sys.argv[1]) if len(sys.argv) > 1 else 2025
    out = Path(f"gastos_deducibles_{year}.xlsx")
    create_template(out, year)
    print(f"Template created: {out}")
