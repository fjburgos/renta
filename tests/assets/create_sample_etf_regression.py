"""
Script to generate sample_etf_transactions.xlsx — a fictional DeGiro Transactions export
for integration regression testing.

Run from the repo root:
    python tests/assets/create_sample_etf_regression.py

Scenarios:
  ISIN IE0091000001 — Alpha ETF: buy 100@20 in 2020, sell 100@30 in 2021 (+996 EUR net)
  ISIN IE0091000002 — Beta ETF:  buy 200@10 in 2021, sell 200@9  in 2022 (-204 EUR net)

No sales after 2022. All values are fictional. No real personal or financial data.

Expected integration test values:
  Year | total_sales | total_purchases | net_result
  2021 |    2998.00  |      2002.00    |    996.00
  2022 |    1798.00  |      2002.00   |   -204.00
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

OUTPUT = Path(__file__).parent / "sample_etf_transactions.xlsx"

HEADERS = [
    "Fecha", "Hora", "Producto", "ISIN",
    "Bolsa de referencia", "Centro de ejecución",
    "Número", "Precio", None,
    "Valor local", None,
    "Valor EUR", "Tipo de cambio", "Comisión AutoFX",
    "Costes de transacción y/o externos EUR",
    "Total EUR", None, "ID Orden", None,
]

# (date, time, product, isin, exchange, venue, qty, price, cur,
#  local_val, local_cur, eur_val, fx_rate, autofx, costs, total_eur,
#  None, order_id, None)
ROWS = [
    # ISIN IE0091000001 — Alpha ETF
    # Buy 100 shares @ 20.00 EUR on 2020-06-01: total = -2000 - 2 commission = -2002
    ("01-06-2020", "09:00", "Alpha UCITS ETF", "IE0091000001",
     "XET", "XETA", 100, 20.00, "EUR", -2000.00, "EUR", -2000.00,
     None, 0.0, -2.0, -2002.0, None, "ORD-2020-0001", None),
    # Sell 100 shares @ 30.00 EUR on 2021-03-15: total = 3000 - 2 commission = 2998
    ("15-03-2021", "09:00", "Alpha UCITS ETF", "IE0091000001",
     "XET", "XETA", -100, 30.00, "EUR", 3000.00, "EUR", 3000.00,
     None, 0.0, -2.0, 2998.0, None, "ORD-2021-0001", None),

    # ISIN IE0091000002 — Beta Bond ETF
    # Buy 200 shares @ 10.00 EUR on 2021-09-01: total = -2000 - 2 commission = -2002
    ("01-09-2021", "09:00", "Beta Bond UCITS ETF", "IE0091000002",
     "XET", "XETA", 200, 10.00, "EUR", -2000.00, "EUR", -2000.00,
     None, 0.0, -2.0, -2002.0, None, "ORD-2021-0002", None),
    # Sell 200 shares @ 9.00 EUR on 2022-04-10: total = 1800 - 2 commission = 1798
    ("10-04-2022", "09:00", "Beta Bond UCITS ETF", "IE0091000002",
     "XET", "XETA", -200, 9.00, "EUR", 1800.00, "EUR", 1800.00,
     None, 0.0, -2.0, 1798.0, None, "ORD-2022-0001", None),
]


def create() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transacciones"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in ROWS:
        ws.append(row)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 16

    wb.save(OUTPUT)
    print(f"Created: {OUTPUT}")


if __name__ == "__main__":
    create()
