"""
Script to generate sample_transactions.xlsx — a fictional DeGiro export for testing.

Run from the repo root:
    python tests/etfs/fixtures/create_sample.py

The generated file covers 4 scenarios:
  ISIN 1 (IE0001000001) — World ETF: FIFO with partial lot consumption
  ISIN 2 (IE0002000002) — Bond ETF:  Simple loss (computable)
  ISIN 3 (IE0003000003) — Small Cap ETF: Loss + wash sale (deferred)
  ISIN 4 (ES0004000004) — Banco Alfa: Stock with split execution (same order_id, same day)

All values are fictional. No real personal or financial data.
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

OUTPUT = Path(__file__).parent / "sample_transactions.xlsx"

# Column headers matching the real DeGiro format.
# Col 16 is blank in data rows; order_id appears at col 17.
HEADERS = [
    "Fecha", "Hora", "Producto", "ISIN",
    "Bolsa de referencia", "Centro de ejecución",
    "Número", "Precio", None,
    "Valor local", None,
    "Valor EUR", "Tipo de cambio", "Comisión AutoFX",
    "Costes de transacción y/o externos EUR",
    "Total EUR", None, "ID Orden", None,
]

# (date, time, product, isin, exchange, venue, qty, price, cur, local_val, local_cur,
#  eur_val, fx_rate, autofx, costs, total_eur, None, order_id, None)
#   0      1      2        3      4        5     6     7     8     9        10
#  11      12     13       14     15       16    17    18
# Each row: (date, time, product, isin, exchange, venue, qty, price, cur,
#            local_val, local_cur, eur_val, fx_rate, autofx, costs, total_eur,
#            None,  order_id,  None)
#             ↑ col 16 blank, matching real DeGiro format
ROWS = [
    # ── ISIN 1: World ETF — two lots, partial sale ───────────────────────────
    # Lot A: 100 shares @ 100.00, total -10002 (incl. 2 EUR commission)
    ("10-01-2024", "09:00", "World Equity UCITS ETF", "IE0001000001",
     "XET", "XETA", 100, 100.00, "EUR", -10000.00, "EUR", -10000.00,
     None, 0.0, -2.0, -10002.0, None, "ORD-2024-0101", None),
    # Lot B: 50 shares @ 105.00, total -5252
    ("15-03-2024", "09:00", "World Equity UCITS ETF", "IE0001000001",
     "XET", "XETA", 50, 105.00, "EUR", -5250.00, "EUR", -5250.00,
     None, 0.0, -2.0, -5252.0, None, "ORD-2024-0102", None),
    # Sale: 80 shares @ 109.00, total +8718 (80×109 − 2 commission)
    ("20-06-2024", "14:30", "World Equity UCITS ETF", "IE0001000001",
     "XET", "XETA", -80, 109.00, "EUR", 8720.00, "EUR", 8720.00,
     None, 0.0, -2.0, 8718.0, None, "ORD-2024-0103", None),

    # ── ISIN 2: Bond ETF — simple loss ──────────────────────────────────────
    # Buy 200 shares @ 100.00, total -20002
    ("01-02-2024", "09:00", "Euro Government Bond UCITS ETF", "IE0002000002",
     "EAM", "XAMS", 200, 100.00, "EUR", -20000.00, "EUR", -20000.00,
     None, 0.0, -2.0, -20002.0, None, "ORD-2024-0201", None),
    # Sell 200 shares @ 92.50, total +18498 (200×92.50 − 2)
    ("05-09-2024", "11:00", "Euro Government Bond UCITS ETF", "IE0002000002",
     "EAM", "XAMS", -200, 92.50, "EUR", 18500.00, "EUR", 18500.00,
     None, 0.0, -2.0, 18498.0, None, "ORD-2024-0202", None),

    # ── ISIN 3: Small Cap ETF — wash sale ───────────────────────────────────
    # Buy 100 shares @ 100.00, total -10002
    ("01-03-2024", "09:00", "European Small Cap UCITS ETF", "IE0003000003",
     "XET", "XETA", 100, 100.00, "EUR", -10000.00, "EUR", -10000.00,
     None, 0.0, -2.0, -10002.0, None, "ORD-2024-0301", None),
    # Sell 100 shares @ 85.00, total +8498 (100×85 − 2) → LOSS -1504
    ("10-07-2024", "14:00", "European Small Cap UCITS ETF", "IE0003000003",
     "XET", "XETA", -100, 85.00, "EUR", 8500.00, "EUR", 8500.00,
     None, 0.0, -2.0, 8498.0, None, "ORD-2024-0302", None),
    # Re-buy 50 shares @ 86.00 on 2024-08-20 (41 days after sale → wash sale!)
    ("20-08-2024", "09:00", "European Small Cap UCITS ETF", "IE0003000003",
     "XET", "XETA", 50, 86.00, "EUR", -4300.00, "EUR", -4300.00,
     None, 0.0, -2.0, -4302.0, None, "ORD-2024-0303", None),

    # ── ISIN 4: Banco Alfa — split execution (same order_id, same day) ───────
    # Execution row 1: 300 shares @ 10.00, no costs
    ("14-05-2024", "09:03", "Banco Alfa SA", "ES0004000004",
     "MAD", "GROW", 300, 10.00, "EUR", -3000.00, "EUR", -3000.00,
     None, 0.0, None, -3000.0, None, "ORD-2024-0999", None),
    # Execution row 2: 200 shares @ 9.90, 3 EUR costs (same order_id!)
    ("14-05-2024", "09:03", "Banco Alfa SA", "ES0004000004",
     "MAD", "BMEX", 200, 9.90, "EUR", -1980.00, "EUR", -1980.00,
     None, 0.0, -3.0, -1983.0, None, "ORD-2024-0999", None),
    # Sell all 500 shares @ 11.50, total +5747 (500×11.50 − 3)
    ("20-11-2024", "15:00", "Banco Alfa SA", "ES0004000004",
     "MAD", "GROW", -500, 11.50, "EUR", 5750.00, "EUR", 5750.00,
     None, 0.0, -3.0, 5747.0, None, "ORD-2024-1500", None),
]


def create() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transacciones"

    # Header row
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    for row in ROWS:
        ws.append(row)

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 16

    wb.save(OUTPUT)
    print(f"Created: {OUTPUT}")


if __name__ == "__main__":
    create()
