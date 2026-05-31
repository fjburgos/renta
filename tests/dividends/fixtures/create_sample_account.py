"""
Script to generate sample_account.xlsx — a fictional DeGiro Account (Estado de cuenta)
export for integration regression testing.

Run from the repo root:
    python tests/dividends/fixtures/create_sample_account.py

Dividend events (EUR, no FX conversion needed):
  2021: ISIN US0099002100 "Synthetic Corp A" — gross EUR 10.00, withheld EUR 1.50
  2022: ISIN US0099002200 "Synthetic Corp B" — gross EUR 20.00, withheld EUR 3.00
  2023: ISIN US0099002300 "Synthetic Corp C" — gross EUR 25.00, withheld EUR 3.75
  2024: ISIN US0099002400 "Synthetic Corp D" — gross EUR 30.00, withheld EUR 4.50

Expected IRPF values:
  Year | Casilla 0029 (gross) | Casilla 0588 (deductible)
  2021 |  10.00               |  1.50
  2022 |  20.00               |  3.00
  2023 |  25.00               |  3.75
  2024 |  30.00               |  4.50

Deductible = min(withheld_eur, 19% of gross_eur) — all within the 19% cap.
No FX conversion rows are needed since all dividends are in EUR.
All values are fictional. No real personal or financial data.
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

OUTPUT = Path(__file__).parent / "sample_account.xlsx"

HEADERS = [
    "Fecha", "Hora", "Fecha valor", "Producto", "ISIN",
    "Descripción", "Tipo", "Variación", None,
    "Saldo", None, "ID Orden",
]

# Each row: (fecha, hora, fecha_valor, producto, isin,
#            descripcion, tipo, moneda, importe,
#            saldo_moneda, saldo, id_orden)
ROWS = [
    # 2021: Synthetic Corp A — gross 10.00, withheld 1.50, net 8.50
    ("15-06-2021", "10:00", "15-06-2021", "Synthetic Corp A", "US0099002100",
     "Dividendo", None, "EUR", 8.50, "EUR", 8.50, None),
    ("15-06-2021", "10:00", "15-06-2021", "Synthetic Corp A", "US0099002100",
     "Retención del dividendo", None, "EUR", -1.50, "EUR", 7.00, None),

    # 2022: Synthetic Corp B — gross 20.00, withheld 3.00, net 17.00
    ("15-06-2022", "10:00", "15-06-2022", "Synthetic Corp B", "US0099002200",
     "Dividendo", None, "EUR", 17.00, "EUR", 17.00, None),
    ("15-06-2022", "10:00", "15-06-2022", "Synthetic Corp B", "US0099002200",
     "Retención del dividendo", None, "EUR", -3.00, "EUR", 14.00, None),

    # 2023: Synthetic Corp C — gross 25.00, withheld 3.75, net 21.25
    ("15-06-2023", "10:00", "15-06-2023", "Synthetic Corp C", "US0099002300",
     "Dividendo", None, "EUR", 21.25, "EUR", 21.25, None),
    ("15-06-2023", "10:00", "15-06-2023", "Synthetic Corp C", "US0099002300",
     "Retención del dividendo", None, "EUR", -3.75, "EUR", 17.50, None),

    # 2024: Synthetic Corp D — gross 30.00, withheld 4.50, net 25.50
    ("15-06-2024", "10:00", "15-06-2024", "Synthetic Corp D", "US0099002400",
     "Dividendo", None, "EUR", 25.50, "EUR", 25.50, None),
    ("15-06-2024", "10:00", "15-06-2024", "Synthetic Corp D", "US0099002400",
     "Retención del dividendo", None, "EUR", -4.50, "EUR", 21.00, None),
]


def create() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuenta"

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in ROWS:
        ws.append(row)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 14

    wb.save(OUTPUT)
    print(f"Created: {OUTPUT}")


if __name__ == "__main__":
    create()
